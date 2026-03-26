# research/views.py
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required

import openpyxl

from .models import Lecturer, ResearchTitle, ResearchRequest, GuidanceSession
from practicum.models import Practicum
from practicum.serializers import PracticumCreateSerializer



# ══════════════════════════════════════════════════════════
# MIXIN
# ══════════════════════════════════════════════════════════

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff


# ══════════════════════════════════════════════════════════
# PUBLIC VIEWS (User-facing)
# ══════════════════════════════════════════════════════════

class LecturerListView(ListView):
    model = Lecturer
    template_name = 'research/lecturer_list.html'
    context_object_name = 'lecturers'

    def get_queryset(self):
        return Lecturer.objects.filter(is_active=True)


class LecturerDetailView(DetailView):
    model = Lecturer
    template_name = 'research/lecturer_detail.html'


class ResearchTitleListView(ListView):
    model = ResearchTitle
    template_name = 'research/title_list.html'
    context_object_name = 'titles'

    def get_queryset(self):
        return ResearchTitle.objects.filter(
            lecturer_id=self.kwargs['pk'],
            is_active=True
        )
# Tambah di bagian atas views.py, setelah imports
import traceback

def safe_json_view(view_func):
    """Decorator: tangkap semua exception → return JSON error, bukan HTML 500"""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            return view_func(request, *args, **kwargs)
        except Exception as e:
            traceback.print_exc()  # tetap print di terminal
            return JsonResponse({
                'ok':  False,
                'msg': f'Server error: {str(e)}',
            }, status=500)
    return wrapper

# ✅ GANTI DENGAN INI:
class LecturerDeleteView(AdminRequiredMixin, View):

    def post(self, request, pk):
        lecturer = get_object_or_404(Lecturer, pk=pk)

        active_requests = ResearchRequest.objects.filter(
            lecturer=lecturer,
            status__in=['pending', 'approved']
        )

        if active_requests.exists():
            return JsonResponse({
                'ok':     False,
                'action': 'suggest_deactivate',
                'msg':    f'Dosen {lecturer.name} masih punya {active_requests.count()} request aktif. Nonaktifkan saja?',
            })

        title_count = lecturer.research_titles.count()
        ResearchRequest.objects.filter(lecturer=lecturer).delete()
        lecturer.research_titles.all().delete()
        name = lecturer.name
        lecturer.delete()

        return JsonResponse({
            'ok':  True,
            'msg': f'Dosen {name} berhasil dihapus'
                   + (f' beserta {title_count} judul payung' if title_count else '') + '.',
        })


class LecturerDeactivateView(AdminRequiredMixin, View):

    def post(self, request, pk):
        lecturer = get_object_or_404(Lecturer, pk=pk)
        lecturer.is_active = False
        lecturer.save()
        return JsonResponse({
            'ok':  True,
            'msg': f'Dosen {lecturer.name} berhasil dinonaktifkan.',
        })

import json
from practicum.models import Practicum

class JadwalJsonView(AdminRequiredMixin, View):
    def get(self, request, pk):
        from practicum.models import Practicum
        jadwal = get_object_or_404(Practicum, pk=pk)
        return JsonResponse({
            'pk':           jadwal.pk,
            'type':         jadwal.type,
            'session_name': jadwal.session_name,
            'lecturer_id':  jadwal.lecturer_id,
            'date':         str(jadwal.date),          
            'room_id':      jadwal.room_id,
            'start_time':   jadwal.start_time.strftime('%H:%M'),
            'end_time':     jadwal.end_time.strftime('%H:%M'),
            'capacity':     jadwal.capacity,
            'is_active':    jadwal.is_active,
            'notes':        jadwal.description or '',   
        })

class JadwalCreateView(AdminRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'ok': False, 'msg': 'Request body tidak valid.'}, status=400)

        from practicum.serializers import PracticumCreateSerializer
        serializer = PracticumCreateSerializer(data=data)
        if serializer.is_valid():
            obj = serializer.save()
            return JsonResponse({'ok': True, 'msg': 'Jadwal berhasil disimpan!', 'id': obj.pk})
        return JsonResponse({'ok': False, 'errors': serializer.errors}, status=400)


class JadwalUpdateView(AdminRequiredMixin, View):
    def post(self, request, pk):
        from practicum.models import Practicum
        jadwal = get_object_or_404(Practicum, pk=pk)
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'ok': False, 'msg': 'Request body tidak valid.'}, status=400)

        from practicum.serializers import PracticumCreateSerializer
        serializer = PracticumCreateSerializer(jadwal, data=data, partial=False)
        if serializer.is_valid():
            obj = serializer.save()
            return JsonResponse({'ok': True, 'msg': 'Jadwal berhasil diperbarui!', 'id': obj.pk})
        return JsonResponse({'ok': False, 'errors': serializer.errors}, status=400)

class JadwalDeleteView(AdminRequiredMixin, View):
    def post(self, request, pk):
        from practicum.models import Practicum
        jadwal = get_object_or_404(Practicum, pk=pk)

        # Cek apakah ada peserta terdaftar
        if jadwal.registrations.filter(
            status__in=['registered', 'waitlist']
        ).exists():
            return JsonResponse({
                'ok': False,
                'msg': 'Jadwal tidak bisa dihapus — masih ada peserta terdaftar.'
            }, status=400)

        name = jadwal.session_name
        jadwal.delete()
        return JsonResponse({'ok': True, 'msg': f'Jadwal {name} berhasil dihapus.'})

from django.db import transaction

class ResearchRequestCreateView(LoginRequiredMixin, CreateView):
    model = ResearchRequest
    fields = ['lecturer', 'research_title', 'request_type', 'thesis_title', 'proposal']
    login_url = '/login/'

    def is_ajax(self):
        return (
            self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            or 'multipart' in self.request.content_type
            or self.request.content_type == 'application/json'
        )

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        # Ambil research_title_id dari POST
        title_id = request.POST.get('research_title')
        if title_id:
            try:
                title = ResearchTitle.objects.select_for_update().get(pk=title_id)
                if title.is_full:
                    if self.is_ajax():
                        return JsonResponse(
                            {'errors': {'research_title': ['Kuota judul payung ini sudah penuh.']}},
                            status=400
                        )
                    messages.error(request, 'Kuota judul payung sudah penuh.')
                    return redirect('/penelitian/')
            except ResearchTitle.DoesNotExist:
                if self.is_ajax():
                    return JsonResponse(
                        {'errors': {'research_title': ['Judul payung tidak ditemukan.']}},
                        status=400
                    )

        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.student = self.request.user
        response = super().form_valid(form)
        if self.is_ajax():
            return JsonResponse({'id': self.object.pk, 'status': 'ok'})
        return response

    def form_invalid(self, form):
        if self.is_ajax():
            return JsonResponse({'errors': form.errors}, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return '/penelitian/#request-saya'



class ResearchRequestDetailView(DetailView):
    model = ResearchRequest
    template_name = 'research/request_detail.html'


# ══════════════════════════════════════════════════════════
# PENELITIAN MAIN PAGE (Module 4 - student view)
# ══════════════════════════════════════════════════════════

class ResearchListView(LoginRequiredMixin, TemplateView):
    template_name = 'penelitianMain.html'
    login_url = '/login/'

    def get_context_data(self, **kwargs):
        ctx  = super().get_context_data(**kwargs)
        user = self.request.user

        lecturers = Lecturer.objects.filter(is_active=True).prefetch_related(
            'research_titles'
        ).order_by('focus', 'name')

        titles = ResearchTitle.objects.filter(is_active=True)

        ctx['lecturers']             = lecturers
        ctx['total_titles']          = titles.count()
        ctx['available_count']       = sum(1 for t in titles if not t.is_full)
        ctx['full_count']            = sum(1 for t in titles if t.is_full)
        ctx['focus_areas']           = Lecturer.objects.filter(is_active=True).values_list('focus', flat=True).distinct()
        ctx['my_requests']           = ResearchRequest.objects.filter(
                                           student=user
                                       ).select_related('research_title', 'lecturer').order_by('-created_at')
        ctx['my_active_request'] = ResearchRequest.objects.filter(
                                            student=user, status__in=['pending', 'approved']
                                        ).first()
        ctx['my_approved_request'] = ResearchRequest.objects.filter(
                                            student=user,
                                            status='approved'
                                        ).first()
        ctx['guidance_sessions'] = GuidanceSession.objects.filter(
                                            request__student=user,
                                            request__status='approved'  
                                        ).select_related('request').order_by('-date') 
        ctx['user_requested_ids'] = list(
                                        ResearchRequest.objects.filter(
                                            student=user,
                                            status__in=['pending', 'approved']  
                                        ).values_list('research_title_id', flat=True)
                                    )
        # Dosen yang nama-nya sudah boleh terlihat (sudah ada request approved)
        ctx['visible_lecturer_ids']  = list(
                                           ResearchRequest.objects.filter(student=user, status='approved')
                                           .values_list('lecturer_id', flat=True)
                                       )
        return ctx


# ══════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════

@staff_member_required
def export_praktikum(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jadwal Praktikum"

    headers = ['No', 'Tipe', 'Nama Sesi', 'Dosen', 'Tanggal',
               'Jam Mulai', 'Jam Selesai', 'Ruangan',
               'Kapasitas', 'Terdaftar', 'Status']
    ws.append(headers)

    for i, p in enumerate(Practicum.objects.select_related('lecturer', 'room'), 1):
        ws.append([
            i,
            p.get_type_display() if hasattr(p, 'get_type_display') else '-',
            getattr(p, 'session_name', p.title),
            p.lecturer.name if hasattr(p, 'lecturer') and p.lecturer else '-',
            p.date.strftime('%d/%m/%Y'),
            p.start_time.strftime('%H:%M') if p.start_time else '-',
            p.end_time.strftime('%H:%M') if p.end_time else '-',
            p.room.name if hasattr(p, 'room') and p.room else '-',
            getattr(p, 'capacity', getattr(p, 'max_participants', '-')),
            p.registrations.count() if hasattr(p, 'registrations') else '-',
            'Aktif' if getattr(p, 'is_active', True) else 'Nonaktif',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="jadwal_praktikum.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_dosen(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dosen & Judul Payung"

    headers = ['No', 'Nama Dosen', 'NIP', 'Bidang Fokus',
               'Email', 'No HP', 'Judul Payung', 'Kuota', 'Terpakai', 'Status']
    ws.append(headers)

    i = 1
    for lec in Lecturer.objects.prefetch_related('research_titles'):
        titles = lec.research_titles.all()
        if titles.exists():
            for title in titles:
                ws.append([
                    i, lec.name, lec.nip or '-',
                    lec.get_focus_display(),
                    lec.email or '-', lec.phone or '-',
                    title.title, title.quota, title.slots_used,
                    'Aktif' if title.is_active else 'Nonaktif',
                ])
                i += 1
        else:
            ws.append([
                i, lec.name, lec.nip or '-',
                lec.get_focus_display(),
                lec.email or '-', lec.phone or '-',
                '(belum ada judul payung)', '-', '-',
                'Aktif' if lec.is_active else 'Nonaktif',
            ])
            i += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dosen_judul_payung.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_penelitian(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Request Penelitian"

    headers = ['No', 'Nama Mahasiswa', 'NIM', 'Judul Skripsi',
               'Dosen', 'Judul Payung', 'Tipe', 'Status',
               'Tanggal Request', 'Tanggal Disetujui']
    ws.append(headers)

    for i, req in enumerate(
        ResearchRequest.objects.select_related('student', 'lecturer', 'research_title'), 1
    ):
        ws.append([
            i,
            req.student.get_full_name(),
            getattr(req.student, 'nim_nip', '-') or '-',
            req.thesis_title,
            req.lecturer.name,
            req.research_title.title if req.research_title else 'Individu',
            req.get_request_type_display(),
            req.get_status_display(),
            req.created_at.strftime('%d/%m/%Y'),
            req.approved_at.strftime('%d/%m/%Y') if req.approved_at else '-',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="request_penelitian.xlsx"'
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════
# ADMIN CRUD — DOSEN
# ══════════════════════════════════════════════════════════

class DosenCreateView(AdminRequiredMixin, CreateView):
    model = Lecturer
    fields = ['name', 'nip', 'focus', 'email', 'phone', 'bio', 'photo', 'is_active']
    success_url = reverse_lazy('admin_panel')

    def get_success_url(self):
        return reverse_lazy('admin_panel') + '?tab=dosen'

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok', 'id': self.object.pk, 'name': self.object.name})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
        errors = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
        messages.error(self.request, f'Gagal menyimpan dosen: {errors}')
        return redirect(str(reverse_lazy('admin_panel')) + '?tab=dosen')


class DosenUpdateView(AdminRequiredMixin, UpdateView):
    model = Lecturer
    fields = ['name', 'nip', 'focus', 'email', 'phone', 'bio', 'photo', 'is_active']

    def get_success_url(self):
        return reverse_lazy('admin_panel') + '?tab=dosen'

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok', 'id': self.object.pk})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
        errors = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
        messages.error(self.request, f'Gagal update dosen: {errors}')
        return redirect(str(reverse_lazy('admin_panel')) + '?tab=dosen')


class DosenDeleteView(AdminRequiredMixin, View):

    def post(self, request, pk):
        lecturer = get_object_or_404(Lecturer, pk=pk)
        
        # Hapus semua relasi dulu, baru dosen
        ResearchRequest.objects.filter(lecturer=lecturer).delete()
        lecturer.research_titles.all().delete()
        lecturer.delete()

        return JsonResponse({'ok': True, 'msg': f'Dosen berhasil dihapus.'})


class DosenDeactivateView(AdminRequiredMixin, View):

    def post(self, request, pk):
        lecturer = get_object_or_404(Lecturer, pk=pk)
        lecturer.is_active = False
        lecturer.save()
        return JsonResponse({
            'ok': True,
            'msg': f'Dosen {lecturer.name} berhasil dinonaktifkan.',
        })


class DosenJsonView(AdminRequiredMixin, View):
    def get(self, request, pk):
        lec = get_object_or_404(Lecturer, pk=pk)
        return JsonResponse({
            'pk':        lec.pk,
            'name':      lec.name,
            'nip':       lec.nip or '',
            'focus':     lec.focus,
            'email':     lec.email or '',
            'phone':     lec.phone or '',
            'bio':       lec.bio or '',
            'is_active': lec.is_active,
            'photo_url': lec.photo.url if lec.photo else '',
        })


# ══════════════════════════════════════════════════════════
# ADMIN CRUD — JUDUL PAYUNG
# ══════════════════════════════════════════════════════════

class JudulCreateView(AdminRequiredMixin, CreateView):
    model = ResearchTitle
    fields = ['lecturer', 'title', 'focus', 'quota', 'description', 'is_active']

    def get_success_url(self):
        return reverse_lazy('admin_panel') + '?tab=dosen'

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok', 'id': self.object.pk})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
        errors = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
        messages.error(self.request, f'Gagal menyimpan judul payung: {errors}')
        return redirect(str(reverse_lazy('admin_panel')) + '?tab=dosen')


class JudulUpdateView(AdminRequiredMixin, UpdateView):
    model = ResearchTitle
    fields = ['lecturer', 'title', 'focus', 'quota', 'description', 'is_active']

    def get_success_url(self):
        return reverse_lazy('admin_panel') + '?tab=dosen'

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'ok', 'id': self.object.pk})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
        errors = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
        messages.error(self.request, f'Gagal update judul: {errors}')
        return redirect(str(reverse_lazy('admin_panel')) + '?tab=dosen')


class JudulDeleteView(AdminRequiredMixin, View):
    def post(self, request, pk):
        title = get_object_or_404(ResearchTitle, pk=pk)
        # Hapus request yang terkait dulu
        ResearchRequest.objects.filter(research_title=title).delete()
        title.delete()
        return JsonResponse({'status': 'ok', 'msg': 'Judul payung berhasil dihapus.'})



class JudulJsonView(AdminRequiredMixin, View):
    def get(self, request, pk):
        t = get_object_or_404(ResearchTitle, pk=pk)
        return JsonResponse({
            'pk':          t.pk,
            'lecturer_id': t.lecturer_id,
            'title':       t.title,
            'focus':       t.focus,
            'quota':       t.quota,
            'description': t.description or '',
            'is_active':   t.is_active,
        })


from django.views import View
from django.http import JsonResponse
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import ResearchRequest  # sesuaikan import model Anda

class ResearchRequestApproveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            req = ResearchRequest.objects.get(pk=pk)
            req.status = 'approved'
            req.approved_at = timezone.now()
            req.approved_by = request.user
            req.save(update_fields=['status', 'approved_at', 'approved_by'])  # ← EXPLICIT SAVE
            return JsonResponse({
                'status': 'ok',
                'approved_at': req.approved_at.strftime('%d %b %Y %H:%M'),
                'request_id': req.pk,  # ← tambah ID untuk debug
                'thesis_title': req.thesis_title[:50],  # ← tampilkan judul
            })
        except ResearchRequest.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Request tidak ditemukan.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)



class ResearchRequestRejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        import json
        try:
            body = json.loads(request.body)
            req = ResearchRequest.objects.get(pk=pk)
            req.status = 'rejected'
            req.admin_notes = body.get('admin_notes', '')
            req.save()
            return JsonResponse({'status': 'ok'})
        except ResearchRequest.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Request tidak ditemukan.'}, status=404)
class JudulMahasiswaView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from .models import ResearchTitle, ResearchRequest
        try:
            title = ResearchTitle.objects.get(pk=pk)
        except ResearchTitle.DoesNotExist:
            return JsonResponse({'mahasiswa': []})

        requests = ResearchRequest.objects.filter(
            research_title=title
        ).select_related('student')

        data = [
            {
                'name':        r.student.get_full_name(),
                'nim':         getattr(r.student, 'nim_nip', ''),
                'thesis_title': r.thesis_title,
                'status':      r.status,
            }
            for r in requests
        ]
        return JsonResponse({'mahasiswa': data})

class TitleSlotsView(View):
    def get(self, request):
        titles = ResearchTitle.objects.filter(is_active=True)
        data = [
            {
                'id':         t.pk,
                'slots_used': t.slots_used,
                'quota':      t.quota,
                'is_full':    t.is_full,
            }
            for t in titles
        ]
        return JsonResponse(data, safe=False)
    
    
from datetime import date, datetime  
class GuidanceSessionCreateView(LoginRequiredMixin, View):
    def post(self, request):
        research_req = ResearchRequest.objects.filter(
            student=request.user,
            status='approved'
        ).first()

        if not research_req:
            return JsonResponse({'status': 'error', 'message': 'Tidak ada request yang disetujui.'}, status=403)

        date_str = request.POST.get('date')
        topic    = request.POST.get('topic', '').strip()

        if not date_str or not topic:
            return JsonResponse({'status': 'error', 'message': 'Tanggal dan topik wajib diisi.'}, status=400)

        session = GuidanceSession.objects.create(
            request    = research_req,
            date       = date_str,           # Django simpan sebagai DateField otomatis
            topic      = topic,
            notes      = request.POST.get('notes', ''),
            attachment = request.FILES.get('attachment'),
        )

        # ← FIX: parse dulu ke date object sebelum strftime
        from datetime import date as date_type
        session_date = session.date if hasattr(session.date, 'strftime') \
                       else datetime.strptime(str(session.date), '%Y-%m-%d').date()

        return JsonResponse({
            'status': 'ok',
            'id':     session.id,
            'topic':  session.topic,
            'date':   session_date.strftime('%d %b %Y'),
            'notes':  session.notes,
        })

class GuidanceSessionUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            session = GuidanceSession.objects.get(pk=pk, request__student=request.user)
            session.date = request.POST.get('date')
            session.topic = request.POST.get('topic')
            session.notes = request.POST.get('notes', '')
            if request.FILES.get('attachment'):
                session.attachment = request.FILES['attachment']
            session.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


class GuidanceSessionDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        session = get_object_or_404(
            GuidanceSession,
            pk=pk,
            request__student=request.user 
        )
        session.delete()
        return JsonResponse({'status': 'ok'})




import csv
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views import View
from django.utils import timezone
from practicum.models import Practicum, PracticumRegistration
from research.models import Lecturer, ResearchTitle, ResearchRequest


@method_decorator(staff_member_required, name='dispatch')
class ExportCSVView(View):

    def get(self, request):
        export_type = request.GET.get('type', 'praktikum')
        date_from   = request.GET.get('date_from', '')
        date_to     = request.GET.get('date_to', '')
        status      = request.GET.get('status', '')

        dispatch = {
            'praktikum' : self._export_praktikum,
            'peserta'   : self._export_peserta,
            'dosen'     : self._export_dosen,
            'penelitian': self._export_penelitian,
        }

        handler = dispatch.get(export_type, self._export_praktikum)
        return handler(request, date_from, date_to, status)

    # ------------------------------------------------------------------
    def _make_response(self, filename):
        ts  = timezone.now().strftime('%Y%m%d_%H%M')
        res = HttpResponse(content_type='text/csv; charset=utf-8')
        res['Content-Disposition'] = f'attachment; filename="{filename}_{ts}.csv"'
        res.write('\ufeff')     
        return res
    
    def _export_peserta(self, request, date_from, date_to, status):
        qs = PracticumRegistration.objects.select_related(
            'practicum__lecturer', 'practicum__room', 'student'
        ).all()
        if date_from:
            qs = qs.filter(practicum__date__gte=date_from)
        if date_to:
            qs = qs.filter(practicum__date__lte=date_to)
        if status:
            qs = qs.filter(status=status)

        res    = self._make_response('peserta_praktikum')
        writer = csv.writer(res)
        writer.writerow([
            'No', 'Nama Sesi', 'Tipe', 'Tanggal Jadwal',
            'Nama Mahasiswa', 'NIM/NIP', 'Email',
            'Status Registrasi', 'Kehadiran (%)',
            'Sertifikat Terbit', 'Tgl Daftar'
        ])
        for i, r in enumerate(qs.order_by('practicum__date', 'created_at'), 1):
            student = r.student
            # ── FIX: nim_nip (dengan underscore), bukan nimnip ───────────
            nim = (
                getattr(student, 'nim_nip',  None)
                or getattr(student, 'nimnip', None)
                or getattr(student, 'nim',    None)
                or '-'
            )
            writer.writerow([
                i,
                r.practicum.session_name,
                r.practicum.get_type_display(),
                r.practicum.date.strftime('%d/%m/%Y'),
                student.get_full_name(),
                nim,
                student.email,
                r.get_status_display(),
                getattr(r, 'attendance_percentage', '-'),
                'Ya' if getattr(r, 'certificate_issued', False) else 'Tidak',
                r.created_at.strftime('%d/%m/%Y %H:%M'),
            ])
        return res


    # ------------------------------------------------------------------
    def _export_praktikum(self, request, date_from, date_to, status):
        qs = Practicum.objects.select_related('lecturer', 'room').all()
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        res    = self._make_response('jadwal_praktikum')
        writer = csv.writer(res)
        writer.writerow(['No', 'Tipe', 'Nama Sesi', 'Dosen', 'Tanggal',
                        'Jam Mulai', 'Jam Selesai', 'Ruangan',
                        'Kapasitas', 'Terdaftar', 'Status'])
        for i, p in enumerate(qs.order_by('date', 'start_time'), 1):
            writer.writerow([
                i,
                p.get_type_display(),
                p.session_name,
                p.lecturer.name,
                p.date.strftime('%d/%m/%Y'),
                p.start_time.strftime('%H:%M'),
                p.end_time.strftime('%H:%M'),
                p.room.name,
                p.capacity,
                p.registered_count,   # property sudah ada di model
                'Aktif' if p.is_active else 'Nonaktif',
                ])
        return res



    # ------------------------------------------------------------------
    def _export_dosen(self, request, date_from, date_to, status):
        qs = Lecturer.objects.prefetch_related('research_titles').all()

        res    = self._make_response('dosen_judul_payung')
        writer = csv.writer(res)
        writer.writerow(['No', 'Nama Dosen', 'NIP', 'Bidang', 'Email',
                         'No. HP', 'Judul Payung', 'Kuota', 'Terisi', 'Status Judul'])
        i = 1
        for d in qs.order_by('name'):
            titles = d.research_titles.all()
            if not titles.exists():
                writer.writerow([i, d.name, d.nip or '-', d.get_focus_display(),
                                 d.email or '-', d.phone or '-',
                                 '-', '-', '-', '-'])
                i += 1
            else:
                for t in titles:
                    writer.writerow([
                        i, d.name, d.nip or '-', d.get_focus_display(),
                        d.email or '-', d.phone or '-',
                        t.title, t.quota, t.slots_used,
                        'Aktif' if t.is_active else 'Nonaktif',
                    ])
                    i += 1
        return res

    # ------------------------------------------------------------------
    def _export_penelitian(self, request, date_from, date_to, status):
        from research.models import ResearchRequest
        
        qs = ResearchRequest.objects.select_related(
            'student', 'lecturer', 'research_title'
        ).all()
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if status:
            qs = qs.filter(status=status)

        res    = self._make_response('request_penelitian')
        writer = csv.writer(res)
        writer.writerow([
            'No', 'Nama Mahasiswa', 'NIM/Username', 'Judul Skripsi/Proposal',
            'Dosen Pembimbing', 'Judul Payung', 'Status',
            'Tgl Request', 'Catatan/Alasan',
        ])

        for i, r in enumerate(qs.order_by('-created_at'), 1):
            # Ambil student info
            student      = r.student
            nama         = student.get_full_name() if hasattr(student, 'get_full_name') else str(student)
            nim          = getattr(student, 'nimnip',   None) \
                        or getattr(student, 'nim',       None) \
                        or getattr(student, 'username',  '-')

            # Ambil judul skripsi/proposal
            judul        = getattr(r, 'thesis_title',    None) \
                        or getattr(r, 'judul',           None) \
                        or getattr(r, 'title',           None) \
                        or '-'

            # Dosen
            dosen        = r.lecturer.name if r.lecturer else '-'

            # Judul payung
            payung       = '-'
            if r.research_title:
                payung   = getattr(r.research_title, 'title', str(r.research_title))

            # Status display
            try:
                status_display = r.get_status_display()
            except Exception:
                status_display = r.status

            # Tanggal
            tgl_request  = r.created_at.strftime('%d/%m/%Y') if r.created_at else '-'

            # Catatan admin
            catatan      = getattr(r, 'admin_notes',        None) \
                        or getattr(r, 'notes',              None) \
                        or getattr(r, 'rejection_reason',   None) \
                        or '-'

            writer.writerow([
                i, nama, nim, judul, dosen, payung,
                status_display, tgl_request, catatan,
            ])

        return res

import csv, io
from datetime import datetime
from django.http import JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required


@method_decorator(staff_member_required, name='dispatch')
class ImportCSVView(View):

    def post(self, request):
        import_type = request.POST.get('type')
        file        = request.FILES.get('file')

        if not file or not file.name.endswith('.csv'):
            return JsonResponse({'ok': False, 'errors': ['File harus berformat .csv']})

        # ✅ Fix 1: fallback encoding
        raw = file.read()
        decoded = None
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                decoded = raw.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if decoded is None:
            return JsonResponse({'ok': False, 'errors': ['File tidak bisa dibaca. Simpan ulang sebagai CSV UTF-8.']})

        reader = csv.DictReader(io.StringIO(decoded))
        errors = []
        count  = 0

        try:
            if import_type == 'praktikum':
                count, errors = self._import_praktikum(reader)
            elif import_type == 'dosen':
                count, errors = self._import_dosen(reader)
            elif import_type == 'peserta':
                count, errors = self._import_peserta(reader)
            else:
                return JsonResponse({'ok': False, 'errors': ['Tipe tidak dikenal.']})
        except Exception as e:
            return JsonResponse({'ok': False, 'errors': [str(e)]})

        if errors and count == 0:
            return JsonResponse({'ok': False, 'errors': errors})
        return JsonResponse({'ok': True, 'count': count, 'errors': errors})

    # ── Helpers ───────────────────────────────────────────────────────

    def _parse_date(self, date_str):
        date_str = date_str.strip().replace('\xa0', '')
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        raise ValueError(f'Format tanggal "{date_str}" tidak dikenal. Gunakan YYYY-MM-DD atau DD/MM/YYYY.')

    def _parse_time(self, raw: str) -> str:
        import re
        raw = raw.strip().replace('\xa0', ' ').strip()
        # Strip detik dulu: "13:00:00 AM" → "13:00 AM", "13:00:00" → "13:00"
        raw = re.sub(r'^(\d{1,2}:\d{2}):\d{2}', r'\1', raw)
        # Strip AM/PM setelah itu: "13:00 AM" → "13:00"
        raw = re.sub(r'\s*(AM|PM)$', '', raw, flags=re.IGNORECASE).strip()

        for fmt in ('%H:%M', '%I:%M'):
            try:
                return datetime.strptime(raw, fmt).strftime('%H:%M')
            except ValueError:
                continue
        raise ValueError(f"Format waktu tidak dikenal: '{raw}'")


    # ── Import Praktikum ──────────────────────────────────────────────
    def _import_praktikum(self, reader):
        from practicum.models import Practicum
        from rooms.models import Room
        from .models import Lecturer
        import re, logging
        logger = logging.getLogger(__name__)
        
        errors, count, skipped = [], 0, 0

        for i, row in enumerate(reader, 2):
            try:
                nama_dosen   = row.get('nama_dosen', '').strip()
                nama_ruangan = row.get('nama_ruangan', '').strip()
                
                logger.warning(f"Baris {i}: dosen='{nama_dosen}' ruangan='{nama_ruangan}' jam_selesai='{row.get('jam_selesai','')}'")

                lecturer = Lecturer.objects.filter(name__iexact=nama_dosen).first()
                if not lecturer:
                    raise ValueError(f'Dosen "{nama_dosen}" tidak ditemukan.')

                room = Room.objects.filter(name__iexact=nama_ruangan).first()
                if not room:
                    raise ValueError(f'Ruangan "{nama_ruangan}" tidak ditemukan.')

                parsed_date  = self._parse_date(row['tanggal'])
                parsed_start = self._parse_time(row['jam_mulai'])
                parsed_end   = self._parse_time(row['jam_selesai'])
                
                logger.warning(f"Baris {i}: parsed OK start={parsed_start} end={parsed_end}")

                from django.db import IntegrityError

                try:
                    obj, created = Practicum.objects.get_or_create(
                        room         = room,
                        date         = parsed_date,
                        start_time   = parsed_start,
                        defaults={
                            'type'       : row['tipe'].strip(),
                            'session_name': row['nama_sesi'].strip(),
                            'lecturer'   : lecturer,
                            'end_time'   : parsed_end,
                            'capacity'   : int(row['kapasitas']),
                            'description': row.get('catatan', '').strip(),
                            'is_active'  : True,
                        }
                    )
                    if created:
                        count += 1
                    else:
                        skipped += 1
                except IntegrityError:
                    raise ValueError(
                        f'Ruangan "{nama_ruangan}" sudah dipakai pada {parsed_date} jam {parsed_start}. Konflik jadwal!'
                    )

            except Exception as e:
                errors.append(f'Baris {i}: {e}')
                logger.warning(f"Baris {i}: ERROR {e}")

        if skipped:
            errors.append(f'ℹ️ {skipped} baris dilewati karena sudah ada (duplikat).')

        return count, errors


    # ── Import Dosen ──────────────────────────────────────────────────

    def _import_dosen(self, reader):
        from .models import Lecturer, ResearchTitle
        errors, count = [], 0

        for i, row in enumerate(reader, 2):
            try:
                dosen, _ = Lecturer.objects.get_or_create(
                    name=row['nama'].strip(),
                    defaults={
                        'nip'  : row.get('nip', '').strip(),
                        'focus': row.get('bidang', '').strip(),
                        'email': row.get('email', '').strip(),
                        'phone': row.get('no_hp', '').strip(),
                        'bio'  : row.get('bio', '').strip(),
                    }
                )
                judul = row.get('judul_payung', '').strip()
                if judul:
                    ResearchTitle.objects.get_or_create(
                        lecturer=dosen,
                        title=judul,
                        defaults={
                            'focus'    : row.get('bidang_judul', dosen.focus).strip(),
                            'quota'    : int(row.get('kuota', 5) or 5),
                            'is_active': True,
                        }
                    )
                count += 1
            except Exception as e:
                errors.append(f'Baris {i}: {e}')
        return count, errors

    # ── Import Peserta ────────────────────────────────────────────────

    def _import_peserta(self, reader):
        from practicum.models import Practicum, PracticumRegistration
        from django.contrib.auth import get_user_model
        User = get_user_model()
        errors, count = [], 0

        for i, row in enumerate(reader, 2):
            try:
                jadwal  = Practicum.objects.get(pk=row['jadwal_id'].strip())
                nim     = row['nim_atau_username'].strip()
                student = (User.objects.filter(nimnip=nim).first()
                        or User.objects.filter(username=nim).first())
                if not student:
                    raise ValueError(f'User "{nim}" tidak ditemukan.')
                PracticumRegistration.objects.get_or_create(
                    practicum=jadwal, student=student
                )
                count += 1
            except Exception as e:
                errors.append(f'Baris {i}: {e}')
        return count, errors


import json
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
@require_POST
def bulk_delete_view(request):
    data      = json.loads(request.body)
    ids       = data.get('ids', [])
    item_type = data.get('type', '')

    if item_type == 'praktikum':
        from practicum.models import Practicum
        deleted, _ = Practicum.objects.filter(pk__in=ids).delete()
    elif item_type == 'dosen':
        from .models import Lecturer
        deleted, _ = Lecturer.objects.filter(pk__in=ids).delete()
    else:
        return JsonResponse({'ok': False, 'error': 'Tipe tidak dikenal'})

    return JsonResponse({'ok': True, 'deleted': deleted})
